from __future__ import annotations

import importlib.util
import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.core.config import settings
from app.integrations.resend_client import ResendClient
from app.services.guardrails import ClientGateResult, clean_agency_name, validate_client_notes
from app.services.text_cleanup import clean_packet_text

ROOT = Path(__file__).resolve().parents[2]
LEGACY_ROOT = ROOT / "legacy_money"
LEGACY_WORKERS = LEGACY_ROOT / "Archive" / "future_workers"
MASTER_PATH = LEGACY_ROOT / "handoff_master_log.xlsx"
GENERATE_SCRIPT = LEGACY_WORKERS / "generate_packets.py"

MASTER_HEADERS = [
    "submission_id",
    "respondent_id",
    "submitted_at",
    "email",
    "client_name",
    "focus",
    "tone",
    "raw_notes",
    "generated_packet",
    "status",
    "sent_at",
    "error_notes",
]

FIELD_TO_HEADER = {
    "Where should we send your handoff packet?": "email",
    "Client or company name": "client_name",
    "What should this focus on?": "focus",
    "Preferred tone for the follow-up email": "tone",
    "Paste your rough client call notes": "raw_notes",
}


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, list):
        return ", ".join(_stringify(v) for v in value if _stringify(v))
    if isinstance(value, dict):
        if "value" in value:
            return _stringify(value["value"])
        if "label" in value:
            return _stringify(value["label"])
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def _field_map(payload: dict[str, Any]) -> dict[str, str]:
    data = payload.get("data") or {}
    fields = data.get("fields") or []
    out: dict[str, str] = {}
    for item in fields:
        label = _stringify(item.get("label") or item.get("title") or item.get("key"))
        if not label:
            continue
        out[label] = _stringify(item.get("value"))
    return out


def _header_map(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value:
            out[str(value)] = col
    return out


def ensure_master_workbook() -> dict[str, Any]:
    MASTER_PATH.parent.mkdir(parents=True, exist_ok=True)
    created = False
    changed = False

    if not MASTER_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Log"
        ws.append(MASTER_HEADERS)
        wb.save(MASTER_PATH)
        wb.close()
        created = True

    wb = load_workbook(MASTER_PATH)
    try:
        if "Master Log" not in wb.sheetnames:
            ws = wb.create_sheet("Master Log")
            ws.append(MASTER_HEADERS)
            changed = True
        else:
            ws = wb["Master Log"]

        headers = _header_map(ws)
        if not headers:
            ws.append(MASTER_HEADERS)
            headers = _header_map(ws)
            changed = True

        for header in MASTER_HEADERS:
            if header not in headers:
                ws.cell(1, ws.max_column + 1).value = header
                changed = True

        if changed:
            wb.save(MASTER_PATH)
    finally:
        wb.close()

    return {
        "status": "ok",
        "created": created,
        "changed": changed,
        "master_path": str(MASTER_PATH),
    }


def _load_module(path: Path, module_name: str):
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load module from {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _extract_client_fields(payload: dict[str, Any]) -> dict[str, str]:
    fields = _field_map(payload)
    return {
        "email": fields.get("Where should we send your handoff packet?", ""),
        "client_name": clean_agency_name(fields.get("Client or company name", "")),
        "focus": fields.get("What should this focus on?", "").strip(),
        "tone": fields.get("Preferred tone for the follow-up email", "").strip(),
        "raw_notes": fields.get("Paste your rough client call notes", "").strip(),
    }


def _send_plaintext_fallback(to_email: str, subject: str, body: str) -> dict[str, Any]:
    html = "<br>".join(
        line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        for line in body.splitlines()
    )
    client = ResendClient()
    return client.send_packet_email(to_email=to_email, subject=subject, html=html)


def _guardrail_email(fields: dict[str, str], gate: ClientGateResult) -> tuple[str, str]:
    client_name = fields.get("client_name") or "this submission"

    if gate.status == "unsafe":
        subject = "I could not generate a business handoff from that submission"
        body = (
            "I received the submission, but I could not generate a normal business handoff from the text that was provided.\n\n"
            "The notes appear to contain personal safety or self-harm language instead of usable business call details.\n\n"
            "If this was a mistake or a test, resend the form with the actual business call notes or transcript.\n\n"
            "If the text reflects an urgent real-world safety concern, contact local emergency services or an appropriate crisis resource immediately.\n\n"
            "No business packet was generated.\n\n"
            "— Alan Operator"
        )
        return subject, body

    subject = "Need more detail before I can generate your handoff packet"
    body = (
        f"I received the submission for {client_name}, but there was not enough usable business call detail to generate a real handoff packet.\n\n"
        "Please resend with the actual call notes or transcript, plus any known decisions, next steps, open questions, or requested follow-up.\n\n"
        "No business packet was generated from the current submission.\n\n"
        "— Alan Operator"
    )
    return subject, body


def _find_or_create_row(payload: dict[str, Any]) -> str:
    ensure_master_workbook()

    data = payload.get("data") or {}
    fields = _field_map(payload)
    submission_id = _stringify(data.get("submissionId")) or _stringify(payload.get("submission_id"))
    respondent_id = _stringify(data.get("respondentId"))
    submitted_at = _stringify(data.get("createdAt")) or datetime.now(timezone.utc).isoformat()

    if not submission_id:
        raise RuntimeError("Tally payload missing submissionId")

    wb = load_workbook(MASTER_PATH)
    ws = wb["Master Log"]
    headers = _header_map(ws)

    missing = [h for h in MASTER_HEADERS if h not in headers]
    if missing:
        raise RuntimeError("Master workbook missing headers: " + ", ".join(missing))

    submission_col = headers["submission_id"]
    for row in range(2, ws.max_row + 1):
        existing = _stringify(ws.cell(row, submission_col).value)
        if existing == submission_id:
            print(f"[IMPORT SKIP] submission_id={submission_id} already exists in {MASTER_PATH}")
            return submission_id

    row_data = {h: "" for h in MASTER_HEADERS}
    row_data["submission_id"] = submission_id
    row_data["respondent_id"] = respondent_id
    row_data["submitted_at"] = submitted_at
    row_data["status"] = "new"

    for tally_label, header in FIELD_TO_HEADER.items():
        row_data[header] = fields.get(tally_label, "")

    target_row = ws.max_row + 1
    for header, value in row_data.items():
        ws.cell(target_row, headers[header]).value = value

    wb.save(MASTER_PATH)
    print(
        f"[IMPORT OK] submission_id={submission_id} row={target_row} "
        f"email={row_data['email'] or 'missing'} client={row_data['client_name'] or 'missing'}"
    )
    return submission_id


def run_importer(payload: dict[str, Any]) -> str:
    return _find_or_create_row(payload)


def _builtin_packet(row_data: dict[str, str]) -> str:
    client_name = row_data.get("client_name") or "your client"
    focus = row_data.get("focus") or "next steps and follow-up"
    tone = row_data.get("tone") or "clear, calm, and direct"
    raw_notes = row_data.get("raw_notes") or ""
    notes_preview = "\n".join(line.strip() for line in raw_notes.splitlines() if line.strip())[:2500]
    return clean_packet_text(
        f"""
Post-Call Handoff Packet - {client_name}

Focus
{focus}

Tone
{tone}

Executive recap
The call notes point to a follow-through need around {focus}. The immediate priority is to turn the messy notes into a clear handoff: what was discussed, what matters, what remains open, and what should happen next.

Next steps
1. Confirm the main outcome from the call.
2. Send the follow-up note below while the conversation is still warm.
3. Move the open questions into the next client touchpoint.
4. Put the CRM-ready update into the account record.

Follow-up draft
Hi,

Thanks for the conversation. Based on the call, the main priority is {focus}. The next move is to confirm the open items, align on ownership, and keep momentum without adding extra back-and-forth.

Here are the next steps I have captured:
- Confirm the desired outcome.
- Resolve the open questions.
- Decide who owns the next action.
- Set the next check-in or delivery point.

If I missed anything important, send it over and I will update the handoff.

CRM-ready update
Call completed. Focus: {focus}. Follow-up needed. Open items and next actions should be confirmed with the client before the next delivery step.

Source notes
{notes_preview}
""".strip()
    )


def _run_builtin_generator() -> tuple[int, int]:
    ensure_master_workbook()
    wb = load_workbook(MASTER_PATH)
    generated_count = 0
    failed_count = 0
    try:
        ws = wb["Master Log"]
        headers = _header_map(ws)
        for row in range(2, ws.max_row + 1):
            status = _stringify(ws.cell(row, headers["status"]).value).lower()
            if status not in {"", "new", "imported"}:
                continue

            row_data = {
                header: _stringify(ws.cell(row, col).value)
                for header, col in headers.items()
            }
            if not row_data.get("raw_notes"):
                ws.cell(row, headers["status"]).value = "error"
                ws.cell(row, headers["error_notes"]).value = "missing raw_notes"
                failed_count += 1
                continue

            ws.cell(row, headers["generated_packet"]).value = _builtin_packet(row_data)
            ws.cell(row, headers["status"]).value = "generated"
            ws.cell(row, headers["error_notes"]).value = ""
            generated_count += 1

        wb.save(MASTER_PATH)
    finally:
        wb.close()

    return generated_count, failed_count


def run_generator() -> dict[str, int | str]:
    if GENERATE_SCRIPT.exists():
        if settings.openai_api_key:
            os.environ["OPENAI_API_KEY"] = settings.openai_api_key
        elif not os.getenv("OPENAI_API_KEY", "").strip():
            raise RuntimeError("OPENAI_API_KEY is missing")

        module = _load_module(GENERATE_SCRIPT, "legacy_generate_packets")
        model = os.getenv("OPENAI_MODEL", "gpt-5.5")
        generated_count, failed_count = module.process_workbook(MASTER_PATH, model)
    else:
        model = "builtin_fallback"
        generated_count, failed_count = _run_builtin_generator()
    result = {
        "generated_count": generated_count,
        "failed_count": failed_count,
        "model": model,
    }
    print(f"[GENERATION] {json.dumps(result, ensure_ascii=False)}")
    return result


def _send_resend_email(to_email: str, subject: str, packet: str) -> dict[str, Any]:
    from_email = settings.from_email_fulfillment
    cleaned_packet = clean_packet_text(packet)

    html = "<br>".join(
        line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        for line in cleaned_packet.splitlines()
    )

    client = ResendClient()
    response = client.send_packet_email(
        to_email=to_email,
        subject=subject,
        html=html,
    )
    print(
        f"[RESEND API] from={from_email} to={to_email} "
        f"subject={subject} response={json.dumps(response, default=str, ensure_ascii=False)}"
    )
    return response


def run_sender(submission_id: str | None = None) -> dict[str, Any]:
    wb = load_workbook(MASTER_PATH)
    ws = wb["Master Log"]
    headers = _header_map(ws)

    col_submission_id = headers["submission_id"]
    col_email = headers["email"]
    col_client_name = headers["client_name"]
    col_generated_packet = headers["generated_packet"]
    col_status = headers["status"]
    col_sent_at = headers["sent_at"]
    col_error_notes = headers["error_notes"]

    sent = 0
    failed = 0
    last_submission_id = None
    failure_messages: list[str] = []

    for row in range(2, ws.max_row + 1):
        row_submission_id = _stringify(ws.cell(row, col_submission_id).value)
        if not row_submission_id:
            continue
        if submission_id and row_submission_id != submission_id:
            continue

        status = _stringify(ws.cell(row, col_status).value).lower()
        if status != "generated":
            print(
                f"[SEND SKIP] submission_id={row_submission_id} row={row} status={status or 'blank'} "
                f"target_submission={submission_id or 'all'}"
            )
            continue

        email = _stringify(ws.cell(row, col_email).value)
        client_name = clean_agency_name(_stringify(ws.cell(row, col_client_name).value))
        packet = _stringify(ws.cell(row, col_generated_packet).value)
        subject = f"Your Post-Call Handoff Packet - {client_name}" if client_name and client_name != "your agency" else "Your Post-Call Handoff Packet"

        if not email:
            message = f"Missing email for submission_id={row_submission_id}"
            ws.cell(row, col_status).value = "failed"
            ws.cell(row, col_error_notes).value = message
            print(f"[SEND FAIL] {message}")
            failure_messages.append(message)
            failed += 1
            continue
        if not packet:
            message = f"Missing generated_packet for submission_id={row_submission_id}"
            ws.cell(row, col_status).value = "failed"
            ws.cell(row, col_error_notes).value = message
            print(f"[SEND FAIL] {message}")
            failure_messages.append(message)
            failed += 1
            continue

        try:
            response = _send_resend_email(email, subject, packet)
            ws.cell(row, col_status).value = "sent"
            ws.cell(row, col_sent_at).value = datetime.now(timezone.utc).isoformat()
            ws.cell(row, col_error_notes).value = ""
            sent += 1
            last_submission_id = row_submission_id
            print(
                f"[SEND OK] submission_id={row_submission_id} row={row} to={email} "
                f"response={json.dumps(response, default=str, ensure_ascii=False)}"
            )
        except Exception as exc:
            message = str(exc)[:500]
            ws.cell(row, col_status).value = "failed"
            ws.cell(row, col_error_notes).value = message
            print(f"[SEND FAIL] submission_id={row_submission_id} row={row} to={email} error={message}")
            failure_messages.append(message)
            failed += 1

    wb.save(MASTER_PATH)
    result = {
        "sent_count": sent,
        "failed_count": failed,
        "submission_id": last_submission_id or submission_id or "",
        "failures": failure_messages,
    }
    print(f"[SENDING] {json.dumps(result, ensure_ascii=False)}")
    return result


def run_digest() -> dict[str, int]:
    wb = load_workbook(MASTER_PATH, read_only=True)
    ws = wb["Master Log"]
    headers = _header_map(ws)
    status_col = headers["status"]
    submission_col = headers["submission_id"]

    counts = {"new": 0, "generated": 0, "sent": 0, "failed": 0, "other": 0}
    for row in range(2, ws.max_row + 1):
        submission_id = _stringify(ws.cell(row, submission_col).value)
        if not submission_id:
            continue
        status = _stringify(ws.cell(row, status_col).value).lower()
        if status in counts:
            counts[status] += 1
        else:
            counts["other"] += 1
    print(f"[DIGEST] {json.dumps(counts, ensure_ascii=False)}")
    return counts


def process_tally_submission(payload: dict[str, Any]) -> dict[str, Any]:
    submission_id = _stringify((payload.get("data") or {}).get("submissionId")) or _stringify(payload.get("submission_id"))
    print(f"[TALLY START] submission_id={submission_id or 'missing'}")
    client_fields = _extract_client_fields(payload)
    gate = validate_client_notes(client_fields.get("raw_notes", ""))

    if gate.status != "ok":
        subject, body = _guardrail_email(client_fields, gate)
        send_result = _send_plaintext_fallback(client_fields.get("email", ""), subject, body)
        result = {
            "submission_id": submission_id,
            "generation": {"generated_count": 0, "failed_count": 0, "model": "guardrail"},
            "sending": {
                "sent_count": 1 if client_fields.get("email") else 0,
                "failed_count": 0 if client_fields.get("email") else 1,
                "submission_id": submission_id,
                "failures": [] if client_fields.get("email") else ["missing email"],
                "guardrail": gate.status,
                "reason": gate.reason,
                "response": send_result if client_fields.get("email") else {},
            },
            "digest": run_digest(),
        }
        print(f"[CLIENT GATE] {json.dumps(result, ensure_ascii=False)}")
        return result

    submission_id = run_importer(payload)
    generation = run_generator()
    sending = run_sender(submission_id=submission_id)
    digest = run_digest()
    result = {
        "submission_id": submission_id,
        "generation": generation,
        "sending": sending,
        "digest": digest,
    }
    print(f"[TALLY DONE] {json.dumps(result, ensure_ascii=False)}")
    return result
